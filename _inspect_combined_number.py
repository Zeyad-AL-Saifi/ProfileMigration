import re
from collections import Counter, defaultdict

import openpyxl

MATCH_PATH = r"E:\FIS\Asala\ProfileMigration\excel\Match MF_CLIENT - Copy (3).xlsx"
COMBINED_PATH = r"E:\FIS\Asala\ProfileMigration\ProfileMigration.Api\bin\Debug\net10.0\Clients Details Combined.xlsx"


def print_headers(path: str) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"=== FILE: {path.split(chr(92))[-1]} ===")
    print("Sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        print(f"--- Sheet: {sn} ---")
        print("Header count:", len([h for h in headers if h is not None]))
        for i, h in enumerate(headers):
            if h is not None:
                print(f"  Col{i + 1}: {repr(h)}")
    wb.close()


def redact(val) -> str:
    if val is None:
        return "NULL"
    if isinstance(val, float) and val == int(val):
        s = str(int(val))
    elif isinstance(val, float):
        s = str(val)
    else:
        s = str(val).strip()
    if not s:
        return "EMPTY"
    if re.fullmatch(r"\d+", s):
        return "D" * min(len(s), 10) + (f"({len(s)} digits)" if len(s) > 10 else "")
    if re.fullmatch(r"\d+\.\d+", s):
        parts = s.split(".")
        return f"{'D' * len(parts[0])}.{'D' * len(parts[1])}"
    return re.sub(r"[A-Za-z0-9\u0600-\u06FF]", "X", s)


def normalize(val):
    if val is None:
        return None
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, float):
        return str(val)
    s = str(val).strip()
    return s or None


def analyze_combined_number(path: str) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Match MF_CLIENT"]
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {h: i for i, h in enumerate(headers) if h}
    col_idx = idx["Combined Number"]
    company_idx = idx["COMPANY"]
    client_idx = idx["CLIENT_ID"]

    stats = {
        "total_rows": 0,
        "valid_client_rows": 0,
        "empty_combined": 0,
        "nonempty_combined": 0,
        "types": Counter(),
        "lengths": Counter(),
        "patterns": Counter(),
        "sample_redacted": [],
        "duplicate_values": Counter(),
        "by_company": defaultdict(lambda: {"rows": 0, "empty": 0, "values": Counter()}),
    }
    value_companies = defaultdict(set)

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["total_rows"] += 1
        client_id = row[client_idx]
        company = row[company_idx]
        val = row[col_idx]

        if client_id is None or (isinstance(client_id, (int, float)) and client_id <= 0):
            continue
        stats["valid_client_rows"] += 1

        company_norm = str(company).strip().upper() if company else ""
        bc = stats["by_company"][company_norm]
        bc["rows"] += 1

        norm = normalize(val)
        if norm is None:
            stats["empty_combined"] += 1
            bc["empty"] += 1
            continue

        stats["nonempty_combined"] += 1
        stats["types"][type(val).__name__] += 1
        stats["lengths"][len(norm)] += 1

        if re.fullmatch(r"\d+", norm):
            pat = f"numeric/{len(norm)}"
        elif re.fullmatch(r"\d+\.\d+", norm):
            pat = f"decimal/{len(norm)}"
        else:
            pat = "other"
        stats["patterns"][pat] += 1
        stats["duplicate_values"][norm] += 1
        bc["values"][norm] += 1
        value_companies[norm].add(company_norm)

        seen_pats = {s[0] for s in stats["sample_redacted"]}
        if len(stats["sample_redacted"]) < 8 and pat not in seen_pats:
            stats["sample_redacted"].append((pat, redact(norm)))

    unique_global = sum(1 for _, c in stats["duplicate_values"].items() if c == 1)
    dup_global = sum(1 for _, c in stats["duplicate_values"].items() if c > 1)
    cross_company = {v: comps for v, comps in value_companies.items() if len(comps) > 1}

    print("=== Combined Number analysis (Match MF_CLIENT) ===")
    print(f"Total data rows scanned: {stats['total_rows']}")
    print(f"Valid CLIENT_ID rows: {stats['valid_client_rows']}")
    print(f"Empty Combined Number: {stats['empty_combined']}")
    print(f"Non-empty Combined Number: {stats['nonempty_combined']}")
    if stats["valid_client_rows"]:
        pct = stats["nonempty_combined"] / stats["valid_client_rows"] * 100
        print(f"Coverage: {pct:.2f}%")
    print("Python types:", dict(stats["types"]))
    print("Length distribution (top):", stats["lengths"].most_common(8))
    print("Pattern distribution:", dict(stats["patterns"]))
    print("Global unique values:", unique_global)
    print("Global duplicate keys (value appears 2+ times):", dup_global)
    print("Values shared across multiple COMPANY values:", len(cross_company))
    for v, comps in list(cross_company.items())[:3]:
        print(
            f"  pattern={redact(v)} companies={sorted(comps)} "
            f"count={stats['duplicate_values'][v]}"
        )
    print("By COMPANY:")
    for comp in sorted(stats["by_company"]):
        bc = stats["by_company"][comp]
        dup_keys = sum(1 for _, c in bc["values"].items() if c > 1)
        with_value = bc["rows"] - bc["empty"]
        print(
            f"  {comp}: rows={bc['rows']}, empty={bc['empty']}, "
            f"with_value={with_value}, distinct_values={len(bc['values'])}, "
            f"dup_keys={dup_keys}"
        )
    print("Redacted samples:")
    for pat, sample in stats["sample_redacted"]:
        print(f"  {pat}: {sample}")

    wb.close()

    wb2 = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws2 = wb2["Match MF_CLIENT"]
    fmt_samples = []
    for r in range(2, ws2.max_row + 1):
        cell = ws2.cell(r, col_idx + 1)
        if cell.value is None:
            continue
        fmt = cell.number_format
        if fmt not in [x[1] for x in fmt_samples]:
            fmt_samples.append((redact(cell.value), fmt, type(cell.value).__name__))
        if len(fmt_samples) >= 5:
            break
    print("Excel number_format samples:")
    for val, fmt, typ in fmt_samples:
        print(f"  value_pattern={val} type={typ} number_format={fmt!r}")
    wb2.close()


def main() -> None:
    print_headers(MATCH_PATH)
    print()
    analyze_combined_number(MATCH_PATH)
    print()
    try:
        print_headers(COMBINED_PATH)
    except FileNotFoundError:
        print(f"Combined workbook not found at: {COMBINED_PATH}")


if __name__ == "__main__":
    main()
