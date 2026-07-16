import re
from collections import Counter, defaultdict

import openpyxl

PATH = r"E:\FIS\Asala\ProfileMigration\excel\Match MF_CLIENT - Copy (3).xlsx"


def normalize(val):
    if val is None:
        return None
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, float):
        return str(val)
    s = str(val).strip()
    return s or None


def mask_digits(s: str) -> str:
    return re.sub(r"\d", "D", s)


wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb["Match MF_CLIENT"]
headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
idx = {h: i for i, h in enumerate(headers) if h}

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    cid = row[idx["CLIENT_ID"]]
    if cid is None or (isinstance(cid, (int, float)) and cid <= 0):
        continue
    company = (str(row[idx["COMPANY"]]).strip().upper() if row[idx["COMPANY"]] else "")
    combined = normalize(row[idx["Combined Number"]])
    national = normalize(row[idx["NATIONAL_ID"]])
    rows.append((company, cid, combined, national))

wb.close()

print("Row count:", len(rows))
print("Combined equals NATIONAL_ID:", sum(1 for _, _, c, n in rows if c == n))
print("Combined ends with CLIENT_ID as suffix:")
for company, cid, combined, _ in rows:
    cid_s = str(int(cid)) if isinstance(cid, float) else str(cid)
    if combined and combined.endswith(cid_s):
        print(f"  YES pattern company={company} client_id_len={len(cid_s)} combined_len={len(combined)}")
        break
else:
    print("  none matched suffix rule")

prefix3 = Counter((c[:3] if c and len(c) >= 3 else c) for _, _, c, _ in rows)
prefix4 = Counter((c[:4] if c and len(c) >= 4 else c) for _, _, c, _ in rows)
print("First-3-char masked prefixes:", {mask_digits(k): v for k, v in prefix3.items()})
print("First-4-char masked prefixes:", {mask_digits(k): v for k, v in prefix4.items()})

by_company_prefix = defaultdict(Counter)
for company, _, combined, _ in rows:
    if combined and len(combined) >= 2:
        by_company_prefix[company][mask_digits(combined[:2])] += 1
print("First-2-digit masked prefix by COMPANY:")
for company in sorted(by_company_prefix):
    print(f"  {company}: {dict(by_company_prefix[company])}")

# formula inspection without values
wb2 = openpyxl.load_workbook(PATH, read_only=False, data_only=False)
ws2 = wb2["Match MF_CLIENT"]
formula_patterns = Counter()
for r in range(2, ws2.max_row + 1):
    cell = ws2.cell(r, idx["Combined Number"] + 1)
    if cell.data_type == "f" and cell.value:
        # mask digits and letters in formula
        f = re.sub(r"\d", "D", str(cell.value))
        f = re.sub(r"[A-Za-z_]+", "WORD", f)
        formula_patterns[f] += 1
    elif cell.value is not None:
        formula_patterns[f"LITERAL:{type(cell.value).__name__}"] += 1
print("Cell storage patterns:", dict(formula_patterns))
wb2.close()

# relation to CLIENT_ID length
client_lens = Counter(len(str(int(x[1]))) for x in rows)
combined_lens = Counter(len(x[2]) for x in rows if x[2])
print("CLIENT_ID length distribution:", dict(client_lens))
print("Combined Number length distribution:", dict(combined_lens))
